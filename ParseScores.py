#Code by Luke Farritor
#Project started on Mar 1, 2016

#Imports
import openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string


#Pramaters
oldWb = "Scores.xlsx"   #Downloaded Wb Name
oldSheetName = "Sheet1" #Sheet to get scores from
oldSheetRows = 2818     #Rows in old sheet to be worked with
oldSheetStartRow = 2    #Row that data starts on
newWb = "Parsed.xlsx"   #New Wb Name



#Loading Scoring Sheet
print("Loading Workbook... (This will take a moment)")
oldscrFl = openpyxl.load_workbook(oldWb)
print("Loading Sheet...")
oldScrSht = oldscrFl.get_sheet_by_name(oldSheetName)
print("Loaded.")

#Make new sheet
print("Making new sheet titled " + newWb + "...")
scrsFl = openpyxl.Workbook()
scrsSht = scrsFl.active
print("Beginning to parse match data")

def scorePosition(pos):
    return {
        '1': 5,  #beacon repair
        '2': 5,  #floor goal
        '3': 5,  #on mountain and floor
        '4': 10, #mountain low
        '5': 20, #mountain mid
        '6': 40, #mountain high
    }.get(pos, 0)


def parse ( row ):
    #get match data
    #red
    r1 = int(oldScrSht['H' + str(row)].value)
    r2 = int(oldScrSht['I' + str(row)].value)
    r3 = int(oldScrSht['J' + str(row)].value)
    r1APos = int(oldScrSht['AA' + str(row)].value)
    r2APos = int(oldScrSht['AB' + str(row)].value)
    r1TPos = int(oldScrSht['AE' + str(row)].value)
    r2TPos = int(oldScrSht['AF' + str(row)].value)
    redScr = 0

    #blue
    b1 = int(oldScrSht['K' + str(row)].value)
    b2 = int(oldScrSht['L' + str(row)].value)
    b3 = int(oldScrSht['M' + str(row)].value)
    b1APos = int(oldScrSht['AS' + str(row)].value)
    b2APos = int(oldScrSht['AT' + str(row)].value)
    b1TPos = int(oldScrSht['AW' + str(row)].value)
    b2TPos = int(oldScrSht['AX' + str(row)].value)
    blueScr = 0

    winner = 't' #t = tie, r = red won, b = blue won

    #get match scores
    #Red
    redScr += scorePosition(str(r1APos)) #position at end of auto
    redScr += scorePosition(str(r2APos)) #position at end of auto
    redScr += int(oldScrSht['AC' + str(row)].value) * 20 #beacon
    redScr += int(oldScrSht['AD' + str(row)].value) * 10 #climbers in shelter
    #Teleop
    redScr += scorePosition(str(r1TPos)) #position at end of teleop
    redScr += scorePosition(str(r2TPos)) #position at end of teleop
    redScr += int(oldScrSht['AG' + str(row)].value) * 1 #Floor Goal
    redScr += int(oldScrSht['AI' + str(row)].value) * 5 #Low Goal
    redScr += int(oldScrSht['AJ' + str(row)].value) * 10 #Mid Goal
    redScr += int(oldScrSht['AH' + str(row)].value) * 15 #High Goal
    redScr += int(oldScrSht['AK' + str(row)].value) * 10 #Climbers in Shelter
    redScr += int(oldScrSht['AL' + str(row)].value) * 20 #Zip Line
    redScr += int(oldScrSht['AM' + str(row)].value) * 20 #All Clear
    redScr += int(oldScrSht['AN' + str(row)].value) * 80 #Pull up

    #Blue
    blueScr += scorePosition(str(b1APos)) #position at end of auto
    blueScr += scorePosition(str(b2APos)) #position at end of auto
    blueScr += int(oldScrSht['AU' + str(row)].value) * 20 #beacon
    blueScr += int(oldScrSht['AV' + str(row)].value) * 10 #climbers in shelter
    #Teleop
    blueScr += scorePosition(str(b1TPos)) #position at end of teleop
    blueScr += scorePosition(str(b2TPos)) #position at end of teleop
    blueScr += int(oldScrSht['AY' + str(row)].value) * 1 #Floor Goal
    blueScr += int(oldScrSht['BA' + str(row)].value) * 5 #Low Goal
    blueScr += int(oldScrSht['BB' + str(row)].value) * 10 #Mid Goal
    blueScr += int(oldScrSht['AZ' + str(row)].value) * 15 #High Goal
    blueScr += int(oldScrSht['BC' + str(row)].value) * 10 #Climbers in Shelter
    blueScr += int(oldScrSht['BD' + str(row)].value) * 20 #Zip Line
    blueScr += int(oldScrSht['BE' + str(row)].value) * 20 #All Clear
    blueScr += int(oldScrSht['BF' + str(row)].value) * 80 #Pull up

    if(blueScr > redScr):
        winner = 'b'
    elif(redScr > blueScr):
        winner = 'r'

    if(r3 == 0):
        r3 = ''
    if(b3 == 0):
        b3 = ''
        
    
    #print(str(r1) + ',' + str(r2) + ',' + str(r3) + ' (' + str(redScr) + ') ' + " vs. " + str(b1) + ', ' + str(b2)+ ',' + str(b3) + ' (' + str(blueScr) + ') ' + 'WINNER: ' + winner)

    scrsSht['A' + str(row)] = oldScrSht['A' + str(row)].value #move date
    scrsSht['B' + str(row)] = oldScrSht['B' + str(row)].value #move event name
    scrsSht['C' + str(row)] = oldScrSht['C' + str(row)].value #move event region
    scrsSht['D' + str(row)] = oldScrSht['D' + str(row)].value #move event type
    
    scrsSht['E' + str(row)] = r1
    scrsSht['F' + str(row)] = r2
    scrsSht['G' + str(row)] = r3

    scrsSht['H' + str(row)] = redScr

    scrsSht['I' + str(row)] = b1
    scrsSht['J' + str(row)] = b2
    scrsSht['K' + str(row)] = b3

    scrsSht['L' + str(row)] = blueScr
    scrsSht['M' + str(row)] = winner

    scrsSht['N' + str(row)] = str(r1) + ',' + str(r2) + ',' + str(r3) + ' (' + str(redScr) + ') ' + " vs. " + str(b1) + ', ' + str(b2)+ ',' + str(b3) + ' (' + str(blueScr) + ') ' + 'WINNER: ' + winner
    

for row in range(oldSheetStartRow, oldSheetRows):
    parse(row)
scrsFl.save('parsed.xlsx')
print('complete')
print('press enter to continue')
t = input()

